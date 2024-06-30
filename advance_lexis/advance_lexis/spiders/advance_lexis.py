import json
import urllib.parse
from datetime import datetime

import requests
from openpyxl import load_workbook
from scrapy import Spider, Request, Selector
#  https://elib.tcd.ie/login?url=https://advance.lexis.com/nexis?identityprofileid=69Q2VF60797
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager


class advance_lexisSpider(Spider):
    name = "advance_lexis"

    custom_settings = {
        'FEED_URI': "advance_lexis.xlsx",
        'FEED_FORMAT': 'xlsx',
        # 'FEED_EXPORT_ENCODING': 'utf-8-sig',
        'FEED_EXPORTERS': {'xlsx': 'scrapy_xlsx.XlsxItemExporter'}
    }

    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'Accept-Language': 'en-US,en;q=0.9',
        'Cache-Control': 'no-cache',
        'Connection': 'keep-alive',
        # 'Cookie': '.AspNet.Wam=c07e3dfe-fa9a-406e-b183-679263c268d7%3Andc; LexisMachineId=d1906809-82c8-4c27-bb70-4e393c849ed5; ezproxy=AeRq1W2lRTkC2xv; ezproxyl=AeRq1W2lRTkC2xv; ezproxyn=AeRq1W2lRTkC2xv; ; lna2=ZTZjZDMzZDVlZmVmNTBiYWRhNjAwNDgxNTk4MjFiM2NhOTRjOTRlYzA3NTYzZDI5Njc2MjNjOTY3ZjI4OTMwMDY2MzhiOTg5dXJuOnVzZXI6UEExODc1MzUyNTIhMTAwMDIwMiwxMDAxMDczLDE1MzkxMDksIW5vbmVe; ASP.NET_SessionId=ba78beea-c39c-497b-b9a8-2448510c5a04; LNPAGELOAD-0c08d2af-598f-4de5-80bb-4e6622d77f50=; LNPAGELOAD-2df671b8-6290-487f-a261-110f0c3232f2=; LNPAGELOAD-341e26c1-8a26-4c0e-a172-fc8bafc2d998=; LNPAGELOAD-b47353c7-5968-4aca-8c83-e00ee6795caf=; LNPAGELOAD-eb0fc810-9c97-4477-a3d6-d1771c3ed66c=; LNPAGELOAD-f4dcede2-cdba-4b18-9081-5f3c767b49d6=; LexisMachineId=32c15f1e-b5c2-4276-abd1-17ee6e232e0d; X-LN-Session-TTL=2024-05-06T16%3A05%3A45Z%2C2024-05-06T13%3A05%3A45Z',
        'Pragma': 'no-cache',
        'Referer': 'https://advance-lexis-com.elib.tcd.ie/firsttime?crid=11dbaa55-99cd-406c-8ba3-891ca1585d40&pdmfid=1519360',
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'same-origin',
        'Sec-Fetch-User': '?1',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
        'sec-ch-ua': '"Chromium";v="124", "Google Chrome";v="124", "Not-A.Brand";v="99"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"'
    }
    nx_url = "https://advance-lexis-com.elib.tcd.ie/r/searchresults/hwmyk/results"
    page_payload = json.dumps({
        "id": "results",
        "props": {
            "action": "NextPage"
        }
    })
    page_headers = {
        'Accept': 'application/json, text/javascript, */*; q=0.01',
        'Accept-Language': 'en-US,en;q=0.9',
        'Cache-Control': 'no-cache',
        'Connection': 'keep-alive',
        'Content-Type': 'application/json',
        'Cookie': '',
        # 'Cookie': 'encmipID=ENCR1AES0004|334C21010B3277509EAD07AFF9E833F3; ezproxy=noaUbAI4IRn51AE; ezproxyl=noaUbAI4IRn51AE; ezproxyn=noaUbAI4IRn51AE; aciCookie=ndc; LNSI=14e9c7d5-3873-4508-8b01-2d2344b22c4e; encUserPermID=ENCR1AES0004|6643921476E345A889E370C6E623D22A; LoginTypeCookie=ENCR1AES0004|AA36E9A455DBA32FA9658E3604E257B7C56057DE1205E3C5F7BE58D03EECF8D2ABABCF4FDE03CD482E984B223A93DDF31B3651908359D311EEE07FEF439598DFADD02DFF3DFF71416EC7FA4F75F38BA2; ASP.NET_SessionId=1ca85035-9683-4cd4-81e4-a69ba4fea09d; X-LN-InitialSignOn=; .AspNet.Wam=14e9c7d5-3873-4508-8b01-2d2344b22c4e%3Andc; LexisMachineId=679b68d9-7310-4b92-a953-256ff15e7119; lna2=ZTc1ODU3NmI4ZjI0YzllNDZkZmE4ZjNjMDBkMzI4ZTQ1MDljN2M3M2M0YzE4NWY5YTEyYjI2YjgzOGFhMGI4NzY2M2NhOTIwdXJuOnVzZXI6UEExODc1MzUyNTUhMTAwMDIwMiwxMDAxMDczLDE1MzkxMDksIW5vbmVe; X-LN-Session-TTL=2024-05-09T15%3A44%3A48Z%2C2024-05-09T12%3A44%3A48Z; Perf=%7B%22name%22%3A%22searchresultlist-_handlerPagination-nextpage%22%2C%22sit%22%3A%221715247968580%22%7D',
        # 'Origin': 'https://advance-lexis-com.elib.tcd.ie',
        'Pragma': 'no-cache',
        # 'Referer': 'https://advance-lexis-com.elib.tcd.ie/search/?pdmfid=1519360&pdsearchterms=Students+and+California+and+Gaza+and+Protest+and+Violence+and+not+Haiti&crid=b1c60b7a-1038-4512-b560-4d64ef318102',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
        'X-LN-PreviousRequestId': '',
        'X-Requested-With': 'XMLHttpRequest',
        'sec-ch-ua': '"Chromium";v="124", "Google Chrome";v="124", "Not-A.Brand";v="99"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"'
    }

    def __init__(self):
        self.rows = []
        workbook = load_workbook(filename="input_file.xlsx", read_only=True, data_only=True, keep_links=False)
        sheet_name = "Sheet1"
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True, min_row=2):
            self.rows.append(list(row))

    def start_requests(self):
        options = Options()
        driver = webdriver.Chrome(options=options,
                                  service=Service(ChromeDriverManager().install()))
        driver.get('https://elib.tcd.ie/login?url=https://advance.lexis.com/nexis?identityprofileid=69Q2VF60797')
        time.sleep(3)
        try:
            driver.find_element(By.CSS_SELECTOR, 'input[name="user"]').send_keys('') # TODO UserName
            time.sleep(5)
            driver.find_element(By.CSS_SELECTOR, 'input[name="pass"]').send_keys('') # TODO Password
            time.sleep(5)
            driver.find_element(By.CSS_SELECTOR, 'input[name="Submit2"]').click()
        except:
            pass
        time.sleep(10)

        cookies = {}
        for cook in driver.get_cookies():
            cookies.update({cook.get('name', ''): f"{cook.get('value', '')};"})
        # cookies.pop('X-LN-Session-TTL')
        cookie_dict = '; '.join(f"{cooki.get('name', '')}={cooki.get('value', '')};" for cooki in driver.get_cookies())

        self.headers['Cookie'] = cookie_dict
        driver.close()

        for row in self.rows[:1]:
            keyword = row[0].replace(' ', '+')
            url = "https://advance-lexis-com.elib.tcd.ie/search/?pdmfid=1519360&pdsearchterms={}&pdtimeline=22%2F02%2F2024+to+25%2F02%2F2024%7Cbetween%7CDD%2FMM%2FYYYY".format(
                keyword)
            yield Request(url, callback=self.articles, headers=self.headers, cookies=cookies,
                          meta={'art_id': 1519360, 'keyword': keyword.replace('+', ' '),
                                'row': row, 'articles': [], 'cookies': cookie_dict})

    def articles(self, response):
        cookies = response.meta.get('cookies')
        total_articles = response.meta['articles']
        keyword = response.meta.get('keyword')
        art_id = response.meta['art_id']
        # cookies_str = response.request.headers.getlist('Cookie')[0].decode()
        # cookies = {}
        # for cook in cookies_str.split(';'):
        #     cookies.update({cook.split('=')[0]: f"{cook.split('=')[-1]};"})
        self.headers['Cookie'] = cookies

        script = \
        response.css('script:contains("componentmodels")').get('{}').split('); this.set')[0].split("page.model',")[-1]
        if script:
            data = json.loads(script)
            articles = \
            data.get('collections', {}).get('componentmodels', {}).get('collections', {}).get('featureproviders', [{}])[
                0].get('collections', {}).get('results', {}).get('collections', {}).get('rows', [])
            crid = data.get('collections').get('componentmodels').get('props').get('crid')

        else:
            json_data = json.loads(response.text)
            articles = json_data.get('collections').get('rows')
            crid = response.headers.get('X-LN-CurrentRequestId')
        row = response.meta['row']
        number_of_art = row[3]
        for art in articles:
            date = datetime.strptime(art.get('props', {}).get('publisheddate', ''), "%B %d, %Y %A")
            start_date = datetime.strptime(row[1], '%m/%d/%y')
            end_date = datetime.strptime(row[2], '%m/%d/%Y')
            if start_date <= date <= end_date:
                if not any(arti.get('props', {}).get('title', '') == art.get('props', {}).get('title', '') for arti in
                           total_articles):
                    total_articles.append(art)

        #  parsing detail pages
        # if len(total_articles) == number_of_art:
        sorted_articles = sorted(total_articles, key=lambda x: x['props']['updated'])
        for art in sorted_articles[:number_of_art]:
            article = {
                'Article name': art.get('props', {}).get('title', ''),
                'Body of Article': '',
                'Press': '',
                'Date posted': art.get('props', {}).get('publisheddate', ''),
                'Byline': art.get('props', {}).get('byline', ''),
                'Length': art.get('props', {}).get('length', ''),
                'Language': art.get('props', {}).get('language', ''),
                'Subject': '',
                'Company': '',
                'Industry': '',
                'Person': '',
                'Load date': '',
                'Source information': '',
                'Keyword': keyword
            }
            doc = urllib.parse.quote(art.get('props', {}).get('docfullpath', ''))
            component_id = art.get('props', {}).get('contentcomponentid', '')
            url = f"https://advance-lexis-com.elib.tcd.ie/document/?pdmfid={response.meta['art_id']}&pddocfullpath={doc}&pdcontentcomponentid={component_id}"
            yield Request(url, callback=self.parse_article, headers=self.headers,
                          meta={'article': article})

        # pagination
        if len(total_articles) < number_of_art:
            self.page_headers['Cookie'] = cookies
            self.page_headers['X-LN-PreviousRequestId'] = crid
            response1 = requests.patch(self.nx_url, headers=self.page_headers, data=self.page_payload)
            yield from self.pagination_articles(response1, total_articles, cookies, row, keyword, art_id)
            # method='PATCH')

    def pagination_articles(self, response, total_articles, cookies, row, keyword, art_id):
        json_data = json.loads(response.text)
        articles = json_data.get('collections').get('rows')
        crid = response.headers.get('X-LN-CurrentRequestId')
        number_of_art = row[3]
        for art in articles:
            date = datetime.strptime(art.get('props', {}).get('publisheddate', ''), "%B %d, %Y %A")
            start_date = datetime.strptime(row[1], '%m/%d/%y')
            end_date = datetime.strptime(row[2], '%m/%d/%Y')
            if start_date <= date <= end_date:
                if not any(arti.get('props', {}).get('title', '') == art.get('props', {}).get('title', '') for arti in
                           total_articles):
                    total_articles.append(art)

        # if len(total_articles) == number_of_art:
        sorted_articles = sorted(total_articles, key=lambda x: x['props']['updated'])
        for art in sorted_articles[:1]:
            article = {
                'Article name': art.get('props', {}).get('title', ''),
                'Body of Article': '',
                'Press': '',
                'Date posted': art.get('props', {}).get('publisheddate', ''),
                'Byline': art.get('props', {}).get('byline', ''),
                'Length': art.get('props', {}).get('length', ''),
                'Language': art.get('props', {}).get('language', ''),
                'Subject': '',
                'Company': '',
                'Industry': '',
                'Person': '',
                'Load date': '',
                'Source information': '',
                'Key word': keyword
            }
            cookie = response.cookies

            # Process the cookies
            detail_cookies = '; '.join([f"{cook.name}={cook.value}" for cook in cookie])
            self.headers['Cookie'] = self.headers['Cookie'] + detail_cookies
            doc = urllib.parse.quote(art.get('props', {}).get('docfullpath', ''))
            component_id = art.get('props', {}).get('contentcomponentid', '')
            url = f"https://advance-lexis-com.elib.tcd.ie/document/?pdmfid={art_id}&pddocfullpath={doc}&pdcontentcomponentid={component_id}"
            yield Request(url=url, headers=self.headers, callback=self.parse_article, meta={'article': article})
            # response2 = requests.get(url, headers=self.headers)
            # yield from self.parse_article(response2, article)

        # pagination
        if len(total_articles) < number_of_art:
            self.page_headers['Cookie'] = cookies
            self.page_headers['X-LN-PreviousRequestId'] = crid
            response = requests.patch(self.nx_url, headers=self.page_headers, data=self.page_payload)
            yield from self.pagination_articles(response, total_articles, cookies, row, keyword, art_id)

    def parse_article(self, response):
        article = response.meta['article']
        # response = Selector(text=response.text)
        article['Body of Article'] = ' '.join(response.css('.SS_LeftAlign p *::text').getall()).strip() if response.css('.SS_LeftAlign p *::text').getall() else ''
        article['Subject'] = ' '.join(response.xpath(
            '//span[contains(text(),"Subject")]/following-sibling::text() | //span[contains(text(),"Subject")]/following-sibling::*//text()').getall()).split(
            'Organization')[0].strip() if response.xpath(
            '//span[contains(text(),"Subject")]/following-sibling::text() | //span[contains(text(),"Subject")]/following-sibling::*//text()').getall() else ''
        article['Company'] = ' '.join(response.xpath(
            '//span[contains(text(),"Organization")]/following-sibling::text() | //span[contains(text(),"Organization")]/following-sibling::*//text()').getall()).split(
            'Industry')[0].strip() if response.xpath(
            '//span[contains(text(),"Organization")]/following-sibling::text() | //span[contains(text(),"Organization")]/following-sibling::*//text()').getall() else ''
        article['Industry'] = ' '.join(response.xpath(
            '//span[contains(text(),"Industry")]/following-sibling::text() | //span[contains(text(),"Industry")]/following-sibling::*//text()').getall()).split(
            'Person')[0].strip() if response.xpath(
            '//span[contains(text(),"Industry")]/following-sibling::text() | //span[contains(text(),"Industry")]/following-sibling::*//text()').getall() else ''
        article['Person'] = ' '.join(response.xpath(
            '//span[contains(text(),"Person")]/following-sibling::text() | //span[contains(text(),"Person")]/following-sibling::*//text()').getall()).split(
            'Geographic')[0].strip() if response.xpath(
            '//span[contains(text(),"Person")]/following-sibling::text() | //span[contains(text(),"Person")]/following-sibling::*//text()').getall() else ''
        article['Load date'] = ' '.join(response.xpath(
            '//span[contains(text(),"Load-Date")]/following-sibling::text() | //span[contains(text(),"Load-Date")]/following-sibling::*//text()').getall()).strip() if response.xpath(
            '//span[contains(text(),"Load-Date")]/following-sibling::text() | //span[contains(text(),"Load-Date")]/following-sibling::*//text()') else ''

        yield article
